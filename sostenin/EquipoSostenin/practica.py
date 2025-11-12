import tabula
import pandas as pd
import os # Necesitamos 'os' para comprobar que el archivo existe
# --- Imports para PASO 3 ---
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

# 1. DEFINIR EL ARCHIVO
nombre_archivo = "05 May.pdf" 

print(f"Paso 1: Leyendo el PDF local '{nombre_archivo}'...")

if not os.path.exists(nombre_archivo):
    print(f"¡Error! No se encontró el archivo '{nombre_archivo}'.")
else:
    # 2. LEER EL PDF (con la configuración ganadora)
    try:
        lista_dfs_crudos = tabula.read_pdf(nombre_archivo, pages=1, stream=True, pandas_options={'header': None})

        if not lista_dfs_crudos:
            print("¡Error! Tabula no encontró tablas en el PDF.")
        else:
            print(f"¡ÉXITO! Se encontraron {len(lista_dfs_crudos)} tablas.")
            
            # Seleccionamos la única tabla que encontró
            df_crudo = lista_dfs_crudos[0]
            
            print("\n--- DATOS CRUDOS ---")
            print(df_crudo)
            
            # --- PASO 2: Limpiando los datos con Pandas ---
            print("\nPaso 2: Limpiando los datos con Pandas...")

            # Copiamos el DataFrame crudo
            df_limpio = df_crudo.copy()

            # Definimos los nombres para TODAS las columnas
            nuevos_nombres_columnas = [
                'Monto Exento',
                'Monto Afecto',
                'I.V.A.',
                'Total Mes',
                'Saldo Anterior',
                'Otros Cargos/Abonos',
                'Total a Pagar'
            ]

            # Asignamos los nuevos nombres
            df_limpio.columns = nuevos_nombres_columnas

            # ¡PERO! Nuestra tabla está "de lado".
            # La queremos vertical (Concepto | Valor).
            # Usamos .transpose() para girarla.
            df_limpio = df_limpio.transpose()

            # Ahora el DataFrame se ve así:
            #                         0
            # Monto Exento       10.285
            # Monto Afecto      813.376
            # ...

            # Reiniciamos el índice para que "Monto Exento", etc., sea una columna
            df_limpio = df_limpio.reset_index()

            # Renombramos las columnas finales para el Excel
            df_limpio.columns = ['Concepto', 'Valor ($)']

            print("\n--- DATOS LIMPIOS (procesados por Pandas) ---")
            print(df_limpio)

            # --- PASO 3: Creando el Excel con OpenPyXL ---
            print("\nPaso 3: Creando el Excel con OpenPyXL...")

            wb = Workbook()
            ws = wb.active
            ws.title = "Totales de Boleta"

            # Pegamos el DataFrame LIMPIO
            rows = dataframe_to_rows(df_limpio, index=False, header=True)

            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Aplicar estilos a la cabecera
            estilo_cabecera = Font(bold=True, color="FFFFFF")
            relleno_cabecera = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid") # Fondo azul

            for cell in ws[1]: # ws[1] es la Fila 1 (cabecera)
                cell.font = estilo_cabecera
                cell.fill = relleno_cabecera
                
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15

            # Guardar el archivo
            nombre_archivo_excel = "Mi_Reporte_Final.xlsx"
            wb.save(nombre_archivo_excel)

            print(f"\n¡LISTO! Se ha creado el archivo: {nombre_archivo_excel}")

    except Exception as e:
        print(f"\n¡Ocurrió un error leyendo el PDF con Tabula!")
        print(f"Error: {e}")